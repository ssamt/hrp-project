'''
2016-1~2022-1까지의 시간표 중 따로 process 파일이 없는 것들은
전부 main에서 맨 아래의 변수만 바꿔서 사용이 가능함
'''

from openpyxl import load_workbook
from collections import defaultdict
import csv
import os

def add_students(ws, d):
    row = ws.max_row
    col = ws.max_column
    for i in range(3, row):
        s = ws.cell(i, 1).value
        if s == '신청수':
            break
        s = s.strip()
        idx = s.rfind('(')
        id = s[:idx]
        name = s[idx+1:-1]
        if id not in d:
            d[id] = [name, set()]
        for j in range(2, col):
            if ws.cell(1, j).value is None:
                if i == 3:
                    print(f'Breaked in column {j}')
                break
            if '_' not in ws.cell(1, j).value:
                lecture = ws.cell(1, j).value.strip()
                idx = lecture.rfind('(')
                if idx != -1:
                    lecture = lecture[:idx]
                if i == 3 and lecture not in lectures:
                    print(f'Not proper name of lecture: {lecture}')
                if i == 3 and idx == -1:
                    print(f'Excluded from adding: {lecture}')
                if idx != -1 and ws.cell(i, j).value == 1:
                    d[id][1].add(lecture)

def add_teachers(ws, d):
    row = ws.max_row
    for i in range(2, row+1):
        if ws.cell(i, 1).value is None:
            break
        name = ws.cell(i, 10).value.strip()
        lecture = ws.cell(i, 7).value.strip()
        d[name][lecture] += 1

def connected(s1, s2):
    return s1[0] == s2[0] and abs(int(s1[1])-int(s2[1])) == 1

def add_lectures(ws, d):
    row = ws.max_row
    for i in range(2, row+1):
        name = ws.cell(i, 7).value
        if name is None:
            print(f'Breaked in row {i}')
            break
        name = name.strip()
        if name not in d:
            class_week = ws.cell(i, 14).value #시수
            class_time = ws.cell(i, 11).value.strip().split('|')
            connect = False
            for j in range(len(class_time)):
                for k in range(j+1, len(class_time)):
                    if connected(class_time[j], class_time[k]):
                        connect = True
            d[name] = [0, class_week, connect]
        d[name][0] += 1

def write_students():
    grade1_sheet = wb[ws_1_name]
    grade23_sheet = wb[ws_23_name]
    add_students(grade1_sheet, students)
    add_students(grade23_sheet, students)
    with open(f'{path}/students.csv', 'w', encoding='utf8', newline='') as student_file:
        student_writer = csv.writer(student_file)
        for id in students:
            student_writer.writerow([id, students[id][0]]+list(students[id][1]))

def write_teachers():
    lecture_sheet = wb[ws_lecture_name]
    add_teachers(lecture_sheet, teachers)
    with open(f'{path}/teachers.csv', 'w', encoding='utf8', newline='') as teacher_file:
        teacher_writer = csv.writer(teacher_file)
        for name in teachers:
            l = [name]
            for lec in teachers[name]:
                l.extend([lec, teachers[name][lec]])
            teacher_writer.writerow(l)

def write_lectures():
    lecture_sheet = wb[ws_lecture_name]
    add_lectures(lecture_sheet, lectures)
    with open(f'{path}/lectures.csv', 'w', encoding='utf8', newline='') as lecture_file:
        lecture_writer = csv.writer(lecture_file)
        for name in lectures:
            lecture_writer.writerow([name]+lectures[name])

semester = '2022-1'
workbook_name = {
    '2017-2':'2017학년도 2학기 시간표 및 수강신청 현황_변경 후.xlsx',
    '2018-1':'2018학년도 1학기 시간표 및 수강신청 현황_수학과 수정.xlsx',
    '2020-1':'2020학년도 1학기 시간표 및 수강신청 현황.xlsx',
    '2020-2':'2020학년도 2학기 시간표 및 수강신청 현황.xlsx',
    '2021-1':'2021학년도 1학기 시간표 및 수강신청 현황_수정.xlsx',
    '2021-2':'2021학년도 2학기 시간표 및 수강신청 현황_정정.xlsx',
    '2022-1':'2022학년도 1학기 시간표 및 수강신청 현황.xlsx',
}
wb = load_workbook(f'KSA시간표/{workbook_name[semester]}')
students = dict()
teachers = defaultdict(lambda: defaultdict(int))
lectures = dict()
ws_1_name = {
    '2017-2':'분반현황(1학년)',
    '2018-1':'분반현황(1학년)',
    '2020-1':'분반현황(1학년)',
    '2020-2':'분반현황(1학년)',
    '2021-1':'분반현황(1학년)',
    '2021-2':'분반현황(1학년)',
    '2022-1':'분반현황(1학년)',
}
ws_1_name = ws_1_name[semester]
ws_23_name = {
    '2017-2':'분반 현황(2,3학년)',
    '2018-1':'분반현황(2,3학년)',
    '2020-1':'분반현황(2,3학년)',
    '2020-2':'분반 현황(2,3학년)',
    '2021-1':'분반 현황(2,3학년)',
    '2021-2':'분반현황(2,3학년)',
    '2022-1':'분반 현황(2,3학년)',
}
ws_23_name = ws_23_name[semester]
ws_lecture_name = '담당교원 및 분반별'
folder = semester
path = f'CSV_files/{folder}'
if not os.path.exists(path):
    os.makedirs(path)
write_teachers()
write_lectures()
write_students()
