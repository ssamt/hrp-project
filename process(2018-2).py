from openpyxl import load_workbook
from collections import defaultdict
import csv
import os

def add_students(ws, d):
    row = ws.max_row
    col = ws.max_column
    for i in range(3, row):
        s = ws.cell(i, 1).value
        if s == '신청수':
            break
        s = s.strip()
        idx = s.rfind('(')
        id = s[:idx]
        name = s[idx+1:-1]
        if id not in d:
            d[id] = [name, set()]
        for j in range(2, col-1):
            lecture = ws.cell(1, j).value.strip()
            idx = lecture.rfind('(')
            idx = lecture.rfind('(', 0, idx)
            if idx != -1:
                lecture = lecture[:idx]
            if i == 3 and lecture not in lectures:
                print(f'Not proper name of lecture: {lecture}')
            if i == 3 and idx == -1:
                print(f'Excluded from adding: {lecture}')
            if idx != -1 and ws.cell(i, j).value == 1:
                d[id][1].add(lecture)

def add_teachers(ws, d):
    row = ws.max_row
    for i in range(2, row+1):
        name = ws.cell(i, 11).value.strip()
        lecture = ws.cell(i, 7).value.strip()
        d[name][lecture] += 1

def connected(s1, s2):
    return s1[0] == s2[0] and abs(int(s1[1])-int(s2[1])) == 1

def add_lectures(ws, d):
    row = ws.max_row
    for i in range(2, row+1):
        name = ws.cell(i, 7).value.strip()
        if name not in d:
            class_time = ws.cell(i, 12).value.strip().split('|')
            class_week = len(class_time)
            connect = False
            for j in range(len(class_time)):
                for k in range(j+1, len(class_time)):
                    if connected(class_time[j], class_time[k]):
                        connect = True
            d[name] = [0, class_week, connect]
        d[name][0] += 1

def write_students():
    grade_sheet = wb[ws_name]
    add_students(grade_sheet, students)
    with open(f'{path}/students.csv', 'w', encoding='utf8', newline='') as student_file:
        student_writer = csv.writer(student_file)
        for id in students:
            student_writer.writerow([id, students[id][0]]+list(students[id][1]))

def write_teachers():
    lecture_sheet = wb[ws_lecture_name]
    add_teachers(lecture_sheet, teachers)
    with open(f'{path}/teachers.csv', 'w', encoding='utf8', newline='') as teacher_file:
        teacher_writer = csv.writer(teacher_file)
        for name in teachers:
            l = [name]
            for lec in teachers[name]:
                l.extend([lec, teachers[name][lec]])
            teacher_writer.writerow(l)

def write_lectures():
    lecture_sheet = wb[ws_lecture_name]
    add_lectures(lecture_sheet, lectures)
    with open(f'{path}/lectures.csv', 'w', encoding='utf8', newline='') as lecture_file:
        lecture_writer = csv.writer(lecture_file)
        for name in lectures:
            lecture_writer.writerow([name]+lectures[name])

wb = load_workbook('KSA시간표/2018학년도 2학기 시간표 및 수강신청 현황.xlsx')
students = dict()
teachers = defaultdict(lambda: defaultdict(int))
lectures = dict()
ws_name = '2018학년도 2학기 3차 수강신청 결과'
ws_lecture_name = '2018-2학기 개설과목별 신청인원'
folder = '2018-2'
path = f'CSV_files/{folder}'
if not os.path.exists(path):
    os.makedirs(path)
write_teachers()
write_lectures()
write_students()
